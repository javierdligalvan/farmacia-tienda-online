"""
Procesador de descuentos CINFA vs NORMON
----------------------------------------
Lee la tarifa CINFA del Excel y extrae los descuentos NORMON de la imagen JPG
(ya procesada con OCR y guardada en ocr_raw_normon.txt).
Empareja productos usando un diccionario INN (inglés) → español para resolver
que NORMON usa nombres genéricos internacionales y CINFA usa nombres en español.
"""

import re
import openpyxl
import polars as pl
from pathlib import Path
from collections import Counter

BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "Tarifa Marta.xlsx"
OCR_PATH   = BASE_DIR / "ocr_raw_normon.txt"   # texto OCR ya disponible
OUT_CSV    = BASE_DIR / "COMPARATIVA_DESCUENTOS.csv"

# ─────────────────────────────────────────────────────────────────
# DICCIONARIO INN (inglés/latín) → variantes en español
# Cubre todos los principios activos encontrados en el OCR de NORMON
# ─────────────────────────────────────────────────────────────────
INN_ES: dict[str, list[str]] = {
    # A
    "ACECLOFENAC":            ["ACECLOFENACO"],
    "ACETYLCYSTEINE":         ["ACETILCISTEINA", "ACETILCISTEÍNA"],
    "ACICLOVIR":              ["ACICLOVIR"],
    "ALENDRONATE":            ["ALENDRONATO", "AC. ALENDRON", "ACIDO ALENDRONICO", "ALENDRONICO"],
    "ALLOPURINOL":            ["ALOPURINOL"],
    "ALMOTRIPTAN":            ["ALMOTRIPTAN"],
    "ALPRAZOLAM":             ["ALPRAZOLAM"],
    "AMBROXOL":               ["AMBROXOL"],
    "AMIKACIN":               ["AMIKACINA"],
    "AMISULPRIDE":            ["AMISULPRIDA"],
    "AMLODIPINE":             ["AMLODIPINO", "AMLODIPINA"],
    "AMOXICILLIN":            ["AMOXICILINA"],
    "AMPICILLIN":             ["AMPICILINA"],
    "ANASTROZOLE":            ["ANASTROZOL"],
    "ARIPIPRAZOLE":           ["ARIPIPRAZOL"],
    "ATOSIBAN":               ["ATOSIBAN"],
    "ATENOLOL":               ["ATENOLOL"],
    "ATORVASTATIN":           ["ATORVASTATINA"],
    "AZITHROMYCIN":           ["AZITROMICINA"],
    # B
    "BETAHISTINE":            ["BETAHISTINA"],
    "BICALUTAMIDE":           ["BICALUTAMIDA"],
    "BILASTINE":              ["BILASTINA"],
    "BISOPROLOL":             ["BISOPROLOL"],
    "BOSENTAN":               ["BOSENTAN"],
    "BROMAZEPAM":             ["BROMAZEPAM"],
    # C
    "CALCIUM FOLINATE":       ["ACIDO FOLINICO", "FOLATO CALCICO", "LEUCOVORIN"],
    "CANDESARTAN":            ["CANDESARTAN", "CANDESARTÁN"],
    "CAPECITABINE":           ["CAPECITABINA"],
    "CAPTOPRIL":              ["CAPTOPRIL"],
    "CARBAMAZEPINE":          ["CARBAMAZEPINA"],
    "CARVEDILOL":             ["CARVEDILOL"],
    "CEFAZOLIN":              ["CEFAZOLINA"],
    "CEFIXIME":               ["CEFIXIMA"],
    "CEFOTAXIME":             ["CEFOTAXIMA"],
    "CEFTAZIDIME":            ["CEFTAZIDIMA"],
    "CEFTRIAXONE":            ["CEFTRIAXONA"],
    "CEFUROXIME":             ["CEFUROXIMA"],
    "CEFALEXIN":              ["CEFALEXINA"],
    "CELECOXIB":              ["CELECOXIB"],
    "CETIRIZINE":             ["CETIRIZINA"],
    "CILOSTAZOL":             ["CILOSTAZOL"],
    "CINITAPRIDE":            ["CINITAPRIDA"],
    "CIPROFLOXACIN":          ["CIPROFLOXACINO"],
    "CISATRACURIUM":          ["CISATRACURIO"],
    "CITALOPRAM":             ["CITALOPRAM"],
    "CLARITHROMYCIN":         ["CLARITROMICINA"],
    "CLINDAMYCIN":            ["CLINDAMICINA"],
    "CLORAZEPATE":            ["CLORAZEPATO"],
    "CLOPIDOGREL":            ["CLOPIDOGREL"],
    "CLOXACILLIN":            ["CLOXACILINA"],
    "COLECALCIFEROL":         ["COLECALCIFEROL", "VITAMINA D3", "VITAMINA D"],
    "CYANOCOBALAMIN":         ["CIANOCOBALAMINA", "VITAMINA B12"],
    # D
    "DABIGATRAN":             ["DABIGATRAN", "DABIGATRÁN"],
    "DEFLAZACORT":            ["DEFLAZACORT"],
    "DESLORATADINE":          ["DESLORATADINA"],
    "DESVENLAFAXINE":         ["DESVENLAFAXINA"],
    "DEXKETOPROFEN":          ["DEXKETOPROFENO"],
    "DIAZEPAM":               ["DIAZEPAM"],
    "DICLOFENAC":             ["DICLOFENACO"],
    "DONEPEZIL":              ["DONEPEZILO"],
    "DOXAZOSIN":              ["DOXAZOSINA"],
    "DOXYCYCLINE":            ["DOXICICLINA"],
    "DULOXETINE":             ["DULOXETINA"],
    "DUTASTERIDE":            ["DUTASTERIDA"],
    # E
    "EBASTINE":               ["EBASTINA"],
    "ENALAPRIL":              ["ENALAPRIL"],
    "ENTECAVIR":              ["ENTECAVIR"],
    "EPLERENONE":             ["EPLERENONA"],
    "ERYTHROMYCIN":           ["ERITROMICINA"],
    "ESCITALOPRAM":           ["ESCITALOPRAM"],
    "ESLICARBAZEPINE":        ["ESLICARBAZEPINA"],
    "ESOMEPRAZOLE":           ["ESOMEPRAZOL"],
    "EXEMESTANE":             ["EXEMESTANO"],
    "EZETIMIBE":              ["EZETIMIBA", "EZETIMIBE"],
    # F
    "FAMCICLOVIR":            ["FAMCICLOVIR"],
    "FAMOTIDINE":             ["FAMOTIDINA"],
    "FEBUXOSTATO":            ["FEBUXOSTAT", "FEBUXOSTATO"],
    "FENTANILO":              ["FENTANILO", "FENTANYL"],
    "FESOTERODINE":           ["FESOTERODINA"],
    "FINASTERIDE":            ["FINASTERIDA"],
    "FLECAINIDE":             ["FLECAINIDA"],
    "FLUCONAZOLE":            ["FLUCONAZOL"],
    "FLUOXETINE":             ["FLUOXETINA"],
    "FLUVASTATIN":            ["FLUVASTATINA"],
    # G
    "GABAPENTIN":             ["GABAPENTINA"],
    "GALANTAMINE":            ["GALANTAMINA"],
    "GENTAMICIN":             ["GENTAMICINA"],
    "GLIMEPIRIDE":            ["GLIMEPIRIDA"],
    "GLICAZIDA":              ["GLICLAZIDA", "GLICAZIDA"],
    "GLUCOSAMINE":            ["GLUCOSAMINA"],
    # I
    "IBANDRONIC ACID":        ["ACIDO IBANDRONICO", "IBANDRONICO", "IBANDRONATO", "IBANDRONATE"],
    "IBUPROFEN":              ["IBUPROFENO"],
    "IMATINIB":               ["IMATINIB"],
    "INDAPAMIDA":             ["INDAPAMIDA"],
    "IRBESARTAN":             ["IRBESARTAN"],
    "ITRACONAZOL":            ["ITRACONAZOL"],
    "IVABRADINE":             ["IVABRADINA"],
    # L
    "LACOSAMIDE":             ["LACOSAMIDA"],
    "LAMIVUDINE":             ["LAMIVUDINA"],
    "LAMOTRIGINE":            ["LAMOTRIGINA"],
    "LANSOPRAZOLE":           ["LANSOPRAZOL"],
    "LEFLUNOMIDE":            ["LEFLUNOMIDA"],
    "LERCANIDIPINE":          ["LERCANIDIPINO"],
    "LETROZOLE":              ["LETROZOL"],
    "LEVETIRACETAM":          ["LEVETIRACETAM"],
    "LEVOCETIRIZINE":         ["LEVOCETIRIZINA"],
    "LEVOFLOXACIN":           ["LEVOFLOXACINO"],
    "LEVOBUPIVACAINE":        ["LEVOBUPIVACAINA"],
    "LINEZOLID":              ["LINEZOLID"],
    "LISINOPRIL":             ["LISINOPRIL"],
    "LORATADINE":             ["LORATADINA"],
    "LORAZEPAM":              ["LORAZEPAM"],
    "LORMETAZEPAM":           ["LORMETAZEPAM"],
    "LOSARTAN":               ["LOSARTAN", "LOSARTÁN"],
    "LOVASTATIN":             ["LOVASTATINA"],
    # M
    "MANIDIPINE":             ["MANIDIPINO"],
    "MELOXICAM":              ["MELOXICAM"],
    "MEMANTINE":              ["MEMANTINA"],
    "METAMIZOLE":             ["METAMIZOL"],
    "METFORMIN":              ["METFORMINA"],
    "METHYLPREDNISOLONE":     ["METILPREDNISOLONA"],
    "METRONIDAZOLE":          ["METRONIDAZOL"],
    "MIDAZOLAM":              ["MIDAZOLAM"],
    "MIRTAZAPINE":            ["MIRTAZAPINA"],
    "MONTELUKAST":            ["MONTELUKAST"],
    "MOXIFLOXACIN":           ["MOXIFLOXACINO"],
    "MYCOPHENOLATE":          ["MICOFENOLATO"],
    # N
    "NAPROXEN":               ["NAPROXENO"],
    "NEBIVOLOL":              ["NEBIVOLOL"],
    "NEVIRAPINE":             ["NEVIRAPINA"],
    "NORFLOXACIN":            ["NORFLOXACINO"],
    # O
    "OLANZAPINE":             ["OLANZAPINA"],
    "OLMESARTAN":             ["OLMESARTAN"],
    "OMEPRAZOLE":             ["OMEPRAZOL"],
    "ONDANSETRON":            ["ONDANSETRON", "ONDANSETRÓN"],
    "OXCARBAZEPINE":          ["OXCARBAZEPINA"],
    # P
    "PANTOPRAZOLE":           ["PANTOPRAZOL"],
    "PARACETAMOL":            ["PARACETAMOL"],
    "PARICALCITOL":           ["PARICALCITOL"],
    "PAROXETINE":             ["PAROXETINA"],
    "PIOGLITAZONE":           ["PIOGLITAZONA"],
    "PITAVASTATIN":           ["PITAVASTATINA"],
    "PRAMIPEXOLE":            ["PRAMIPEXOL"],
    "PRAVASTATIN":            ["PRAVASTATINA"],
    "PREGABALIN":             ["PREGABALINA"],
    # Q
    "QUETIAPINE":             ["QUETIAPINA"],
    "QUINAPRIL":              ["QUINAPRIL"],
    # R
    "RABEPRAZOL":             ["RABEPRAZOL"],
    "RAMIPRIL":               ["RAMIPRIL"],
    "RANOLAZINE":             ["RANOLAZINA"],
    "RASAGILINE":             ["RASAGILINA"],
    "REMIFENTANIL":           ["REMIFENTANILO"],
    "REPAGLINIDE":            ["REPAGLINIDA"],
    "RIBAVIRIN":              ["RIBAVIRINA"],
    "RISEDRONATE":            ["RISEDRONATO"],
    "RISPERIDONE":            ["RISPERIDONA"],
    "RIVASTIGMINE":           ["RIVASTIGMINA"],
    "RIZATRIPTAN":            ["RIZATRIPTAN"],
    "ROSUVASTATIN":           ["ROSUVASTATINA"],
    "RUPATADINE":             ["RUPATADINA"],
    # S
    "SERTRALINE":             ["SERTRALINA"],
    "SILDENAFIL":             ["SILDENAFILO"],
    "SILODOSIN":              ["SILODOSINA"],
    "SIMVASTATIN":            ["SIMVASTATINA"],
    "SITAGLIPTIN":            ["SITAGLIPTINA"],
    "SOLIFENACINA":           ["SOLIFENACINA"],
    # T
    "TAMSULOSIN":             ["TAMSULOSINA"],
    "TELMISARTAN":            ["TELMISARTAN"],
    "TERAZOSIN":              ["TERAZOSINA"],
    "TERBINAFINE":            ["TERBINAFINA"],
    "TICAGRELOR":             ["TICAGRELOR"],
    "TOBRAMYCIN":             ["TOBRAMICINA"],
    "TOLTERODINA":            ["TOLTERODINA"],
    "TOPIRAMATE":             ["TOPIRAMATO"],
    "TORASEMIDE":             ["TORASEMIDA", "TORASEMIDE"],
    "TRAMADOL":               ["TRAMADOL"],
    "TRAZODONE":              ["TRAZODONA"],
    # V
    "VALGANCICLOVIR":         ["VALGANCICLOVIR"],
    "VALSARTAN":              ["VALSARTAN"],
    "VARENICLINA":            ["VARENICLINA", "VARENIKLIN"],
    "VENLAFAXINE":            ["VENLAFAXINA"],
    "VILDAGLIPTIN":           ["VILDAGLIPTINA"],
    "VORICONAZOLE":           ["VORICONAZOL"],
    # Z
    "ZOLMITRIPTAN":           ["ZOLMITRIPTAN"],
    "ZOLPIDEM":               ["ZOLPIDEM"],
    "ZONISAMIDA":             ["ZONISAMIDA"],
    # Combinaciones frecuentes (el OCR las muestra como una sola entrada)
    "AMLODIPINE/VALSARTAN":          ["AMLODIPINO/VALSARTAN", "AMLODIPINO VALSARTAN"],
    "AMLODIPINE/ALSARTAN/HIDROCL":   ["AMLODIPINO/OLMESARTAN/HIDROCLOROTIAZIDA"],
    "AMLODIPINO/ATORVASTATINA":      ["AMLODIPINO/ATORVASTATINA", "AMLODIPINO ATORVASTATINA"],
    "AMOXICILLIN/CLAVULANIC":        ["AMOXICILINA/CLAVULANICO", "AMOXICILINA/AC.CLAVULANICO",
                                      "AMOXICILINA ACIDO CLAVULANICO", "AMOXICILINA/CLAVULANICO"],
    "CANDESARTAN/HIDROCLOROTIAZIDA": ["CANDESARTAN/HIDROCLOROTIAZIDA"],
    "CAPTOPRIL HIDROCLOROTIAZIDA":   ["CAPTOPRIL/HIDROCLOROTIAZIDA", "CAPTOPRIL HIDROCLOROTIAZIDA"],
    "DUTASTERIDE/TAMSULOSIN":        ["DUTASTERIDA/TAMSULOSINA", "DUTASTERIDA TAMSULOSINA"],
    "ENALAPRIL/HIDROCLOROTIAZIDA":   ["ENALAPRIL/HIDROCLOROTIAZIDA", "ENALAPRIL HIDROCLOROTIAZIDA"],
    "EZETIMIBE/SIMVASTATIN":         ["EZETIMIBA/SIMVASTATINA", "EZETIMIBA SIMVASTATINA"],
    "EZETIMIBA/ATORVASTATINA":       ["EZETIMIBA/ATORVASTATINA", "EZETIMIBA ATORVASTATINA"],
    "IRBESARTAN/HIDROCLOROTIAZIDA":  ["IRBESARTAN/HIDROCLOROTIAZIDA"],
    "LOSARTAN/HIDROCLOROTIAZIDA":    ["LOSARTAN/HIDROCLOROTIAZIDA"],
    "OLMESARTAN/AMLODIPINO":         ["OLMESARTAN/AMLODIPINO", "OLMESARTAN AMLODIPINO"],
    "OLMESARTAN/HIDROCLOROTIAZIDA":  ["OLMESARTAN/HIDROCLOROTIAZIDA"],
    "OLMESARTAN/AMLODIPINO/HCTZ":    ["OLMESARTAN/AMLODIPINO/HIDROCLOROTIAZIDA"],
    "ROSUVASTATINA/EZETIMIBA":       ["ROSUVASTATINA/EZETIMIBA", "ROSUVASTATINA EZETIMIBA"],
    "SITAGLIPTIN/METFORMIN":         ["SITAGLIPTINA/METFORMINA", "SITAGLIPTINA METFORMINA"],
    "SOLIFENACINA/TAMSULOSINA":      ["SOLIFENACINA/TAMSULOSINA", "SOLIFENACINA TAMSULOSINA"],
    "TELMISARTAN/HIDROCLOROTIAZIDA": ["TELMISARTAN/HIDROCLOROTIAZIDA"],
    "TRAMADOL/PARACETAMOL":          ["TRAMADOL/PARACETAMOL", "TRAMADOL PARACETAMOL"],
    "VALSARTAN/HCZ":                 ["VALSARTAN/HIDROCLOROTIAZIDA"],
    "VILDAGLIPTIN/METFORMIN":        ["VILDAGLIPTINA/METFORMINA", "VILDAGLIPTINA METFORMINA"],
}

# ─────────────────────────────────────────────────────────────────
# UTILIDADES DE NORMALIZACIÓN
# ─────────────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    """Elimina tildes, pasa a mayúsculas y colapsa espacios."""
    s = s.upper().strip()
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ü","U"),("Ñ","N"),
                 ("À","A"),("È","E"),("Ì","I"),("Ò","O"),("Ù","U")]:
        s = s.replace(a, b)
    s = re.sub(r'\s+', ' ', s)
    return s


def _tokens(s: str) -> set[str]:
    """Extrae tokens alfanuméricos significativos (longitud ≥ 3)."""
    return {t for t in re.findall(r'[A-Z0-9]+', _norm(s)) if len(t) >= 3}


# Índice invertido: variante_español_normalizada → INN (se construye tras _norm)
_ES_INDEX: dict[str, str] = {
    _norm(v): inn
    for inn, variantes in INN_ES.items()
    for v in variantes
}


# ─────────────────────────────────────────────────────────────────
# 1. LECTURA DEL EXCEL CINFA
# ─────────────────────────────────────────────────────────────────
def leer_excel_cinfa(path: Path) -> pl.DataFrame:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo, articulo, dto_cinfa = row[0], row[1], row[2]
        if codigo and articulo:
            filas.append({
                "CODIGO":    str(codigo).strip(),
                "NOMBRE":    str(articulo).strip(),
                "DTO_CINFA": int(dto_cinfa) if dto_cinfa is not None else None,
            })
    df = pl.DataFrame(filas, schema={
        "CODIGO":    pl.Utf8,
        "NOMBRE":    pl.Utf8,
        "DTO_CINFA": pl.Int64,
    })
    print(f"✔ Excel CINFA leído: {len(df)} productos")
    return df


# ─────────────────────────────────────────────────────────────────
# 2. PARSEO DEL TEXTO OCR → {inn_normalizado: descuento}
# ─────────────────────────────────────────────────────────────────
# Líneas de ruido a ignorar (no son nombres de principios activos)
_RUIDO = re.compile(
    r'(NORMON|TARIFA|DESCUENTO|LISTA|PRECIO|LABORATORIO|SIN\s+DESCUENTO'
    r'|\d+\s*/\s*\d+\s*MG\s+COMP'      # "875/125 mg comp"
    r'|\d+\s+COMPRIMIDOS?\s*\()'        # "56 comprimidos (lanzamiento)"
    r'|CAPSULAS?\s*$'                   # línea solo con "cápsulas"
    r'|COMPRIMIDOS?\s*$'                # línea solo con "comprimidos"
    r'|SOBRES?\s*$',                    # línea solo con "sobres"
    re.I
)


def parsear_ocr(ocr_text: str) -> dict[str, int]:
    """
    Parsea el texto OCR de NORMON.
    Detecta bloques:  XX%  →  lista de principios activos
    Devuelve {INN_NORMALIZADO: descuento_int}
    'SIN DESCUENTO' se codifica como 0.
    """
    normon_map: dict[str, int] = {}
    descuento_actual: int | None = None
    es_sin_descuento = False

    for linea in ocr_text.splitlines():
        linea = linea.strip()
        if not linea:
            continue

        # ¿Bloque "SIN DESCUENTO"?
        if re.search(r'SIN\s+DESCUENTO', linea, re.I):
            descuento_actual = 0
            es_sin_descuento = True
            continue

        # ¿Línea de porcentaje?
        m_pct = re.match(r'^(\d{1,3})\s*%\s*$', linea)
        if m_pct:
            descuento_actual = int(m_pct.group(1))
            es_sin_descuento = False
            continue

        # ¿Línea de ruido?
        if _RUIDO.search(linea):
            continue

        # Si hay descuento activo y la línea tiene al menos 4 caracteres → principio activo
        if descuento_actual is not None and len(linea) >= 4:
            inn = _norm(linea)
            normon_map[inn] = descuento_actual

    print(f"✔ Principios activos NORMON parseados: {len(normon_map)}")
    resumen = Counter(normon_map.values())
    for pct, n in sorted(resumen.items()):
        label = "SIN DTO" if pct == 0 else f"{pct}%"
        print(f"   {label:8s} →  {n} productos")
    return normon_map


# ─────────────────────────────────────────────────────────────────
# 3. LOOKUP: nombre CINFA → descuento NORMON
# ─────────────────────────────────────────────────────────────────
def _extraer_principio_cinfa(nombre: str) -> str:
    """
    Extrae el principio activo de un nombre CINFA.
    Ejemplo: 'ACECLOFENACO 100 20C' → 'ACECLOFENACO'
             'ENALAPRIL/HIDROCLOROTIAZIDA 20/12.5 20C' → 'ENALAPRIL/HIDROCLOROTIAZIDA'
    Estrategia: toma la parte antes de los primeros dígitos que indiquen dosis.
    """
    nombre_n = _norm(nombre)
    # Quitar sufijo de laboratorio (CINFAMED, NORMON, CINFA…)
    nombre_n = re.sub(r'\s+(CINFAMED|CINFA|NORMON|EFG|MEDE|STADA)\b.*', '', nombre_n)
    # Cortar en el primer token numérico que parezca dosis (p.ej. "100", "20MG", "10/5")
    m = re.search(r'\s+\d[\d.,/]*\s*(MG|MCG|ML|G|UI|IU|%)?\b', nombre_n)
    if m:
        nombre_n = nombre_n[:m.start()].strip()
    return nombre_n


def buscar_dto_normon(nombre_cinfa: str, normon_map: dict[str, int]) -> int | None:
    """
    Busca el descuento NORMON para un nombre CINFA mediante 3 estrategias:
      1. Búsqueda directa del principio activo en el mapa NORMON (normalizado)
      2. Búsqueda en el índice invertido ES→INN y luego en el mapa
      3. Matching por tokens Jaccard ≥ 0.6 contra las claves del mapa
    """
    principio = _extraer_principio_cinfa(nombre_cinfa)

    # ── Estrategia 1: coincidencia directa en mapa NORMON ──────────
    if principio in normon_map:
        dto = normon_map[principio]
        return None if dto == 0 else dto  # 0 = SIN DESCUENTO → devolvemos None

    # ── Estrategia 2: índice invertido español → INN ───────────────
    inn = _ES_INDEX.get(principio)
    if inn:
        inn_norm = _norm(inn)
        if inn_norm in normon_map:
            dto = normon_map[inn_norm]
            return None if dto == 0 else dto

    # ── Estrategia 3: Jaccard sobre tokens del principio activo ────
    toks_cinfa = _tokens(principio)
    if not toks_cinfa:
        return None

    mejor_score = 0.0
    mejor_dto: int | None = None
    for nn, dto in normon_map.items():
        toks_n = _tokens(nn)
        if not toks_n:
            continue
        interseccion = len(toks_cinfa & toks_n)
        union        = len(toks_cinfa | toks_n)
        score = interseccion / union if union else 0
        if score > mejor_score:
            mejor_score = score
            mejor_dto   = dto

    if mejor_score >= 0.60:
        return None if mejor_dto == 0 else mejor_dto
    return None


# ─────────────────────────────────────────────────────────────────
# 4. GENERAR CSV COMPARATIVO
# ─────────────────────────────────────────────────────────────────
def generar_csv(df_cinfa: pl.DataFrame, normon_map: dict[str, int]) -> None:
    dtos_normon = [
        buscar_dto_normon(nombre, normon_map)
        for nombre in df_cinfa["NOMBRE"].to_list()
    ]

    df_final = df_cinfa.with_columns(
        pl.Series("DTO_NORMON", dtos_normon, dtype=pl.Int64)
    ).select(["CODIGO", "NOMBRE", "DTO_CINFA", "DTO_NORMON"])

    df_final.write_csv(OUT_CSV, separator=";")

    n_normon  = df_final["DTO_NORMON"].drop_nulls().len()
    matched   = df_final.filter(pl.col("DTO_CINFA").is_not_null() & pl.col("DTO_NORMON").is_not_null())
    print(f"\n✅ CSV generado: {OUT_CSV}")
    print(f"   Total productos:              {len(df_final)}")
    print(f"   Con descuento CINFA:          {df_final['DTO_CINFA'].drop_nulls().len()}")
    print(f"   Con descuento NORMON:         {n_normon}")
    print(f"   Con AMBOS descuentos:         {len(matched)}")
    if len(matched):
        mejor_normon = matched.filter(pl.col("DTO_NORMON") > pl.col("DTO_CINFA"))
        mejor_cinfa  = matched.filter(pl.col("DTO_CINFA")  > pl.col("DTO_NORMON"))
        iguales      = matched.filter(pl.col("DTO_CINFA")  == pl.col("DTO_NORMON"))
        print(f"   NORMON más ventajoso:         {len(mejor_normon)}")
        print(f"   CINFA  más ventajoso:         {len(mejor_cinfa)}")
        print(f"   Mismo descuento:              {len(iguales)}")
        print()
        print("   TOP 10 diferencias a favor de NORMON:")
        top = (matched
               .with_columns((pl.col("DTO_NORMON") - pl.col("DTO_CINFA")).alias("DIFF"))
               .sort("DIFF", descending=True)
               .head(10))
        for row in top.iter_rows(named=True):
            print(f"     {row['NOMBRE']:<45} CINFA={row['DTO_CINFA']}%  NORMON={row['DTO_NORMON']}%  (+{row['DIFF']}%)")


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    df_cinfa   = leer_excel_cinfa(EXCEL_PATH)
    ocr_texto  = OCR_PATH.read_text(encoding="utf-8")
    normon_map = parsear_ocr(ocr_texto)
    generar_csv(df_cinfa, normon_map)
